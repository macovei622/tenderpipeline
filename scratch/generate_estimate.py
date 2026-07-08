import openpyxl

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Відомість ресурсів"

# Шапка
ws.cell(row=1, column=1, value="ПІДСУМКОВА ВІДОМІСТЬ РЕСУРСІВ (Тестова Вінниця)")
ws.cell(row=3, column=1, value="№")
ws.cell(row=3, column=2, value="Шифр ресурсу")
ws.cell(row=3, column=3, value="Найменування ресурсу")
ws.cell(row=3, column=4, value="Одиниця виміру")
ws.cell(row=3, column=5, value="Кількість")
ws.cell(row=3, column=6, value="Відпускна ціна")
ws.cell(row=3, column=7, value="Кошторисна ціна")
ws.cell(row=3, column=8, value="Кошторисна вартість, грн")

# Розділ
ws.cell(row=4, column=3, value="I. БУДІВЕЛЬНІ МАТЕРІАЛИ")

materials = [
    ("1", "С111-12", "Пісок будівельний", "м3", 200.0, 350.0, 400.0, 80000.0),
    ("2", "С311-45", "Портландцемент М400", "т", 50.0, 4000.0, 4500.0, 225000.0),
    ("3", "С412-10", "Плитка тротуарна ФЕМ", "м2", 1500.0, 600.0, 650.0, 975000.0),
    ("4", "С511-01", "Арматура сталева А500С", "т", 12.0, 34000.0, 36000.0, 432000.0),
    ("5", "С620-05", "Бетон готовий В25 Р4", "м3", 450.0, 2500.0, 2800.0, 1260000.0),
    ("6", "С710-12", "Щебінь фракції 20-40", "м3", 300.0, 500.0, 600.0, 180000.0)
]

for idx, m in enumerate(materials, 5):
    ws.cell(row=idx, column=1, value=m[0])
    ws.cell(row=idx, column=2, value=m[1])
    ws.cell(row=idx, column=3, value=m[2])
    ws.cell(row=idx, column=4, value=m[3])
    ws.cell(row=idx, column=5, value=m[4])
    ws.cell(row=idx, column=6, value=m[5])
    ws.cell(row=idx, column=7, value=m[6])
    ws.cell(row=idx, column=8, value=m[7])

ws.cell(row=11, column=3, value="Разом матеріальних ресурсів")
ws.cell(row=11, column=8, value=3152000.0)

ws.cell(row=12, column=3, value="II. БУДІВЕЛЬНІ МАШИНИ ТА МЕХАНІЗМИ")
ws.cell(row=13, column=1, value="1")
ws.cell(row=13, column=2, value="М100-22")
ws.cell(row=13, column=3, value="Екскаватор одноковшевий")
ws.cell(row=13, column=4, value="маш-год")
ws.cell(row=13, column=5, value=40.0)
ws.cell(row=13, column=7, value=950.0)
ws.cell(row=13, column=8, value=38000.0)

wb.save("test_estimate_winnytsia.xlsx")
wb.close()
print("Test estimate generated successfully!")
