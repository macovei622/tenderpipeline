"""
ai/estimate_parser.py — Парсер кошторисних вивантажень Excel з АВК-5

Цей модуль дозволяє автоматично зчитувати "Підсумкову відомість ресурсів" або
"Локальний кошторис", експортований з АВК-5 в Excel, витягувати звідти список
матеріалів, обсяги та ціни, і передавати в AI-Калькулятор для точного аудиту.
"""
import os
from typing import Optional
from loguru import logger
import openpyxl


def parse_avk_excel_estimate(file_path: str, return_dict: bool = False) -> list[dict] | dict[str, list[dict]]:
    """
    Парсить Excel-файл відомості ресурсів з АВК-5.
    
    Якщо return_dict є True:
        повертає {"materials": [...], "machinery": [...]}
    Інакше:
        повертає список словників матеріалів (materials)
    """
    if not os.path.exists(file_path):
        logger.error(f"Файл кошторису не знайдено: {file_path}")
        return {"materials": [], "machinery": []} if return_dict else []

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    except Exception as e:
        logger.error(f"Не вдалося відкрити Excel файл кошторису: {e}")
        return {"materials": [], "machinery": []} if return_dict else []

    # Беремо перший активний лист
    sheet = wb.active
    logger.info(f"📊 Парсинг кошторису Excel: лист '{sheet.title}', всього рядків: {sheet.max_row or 'невідомо'}")

    materials = []
    machinery = []
    header_found = False
    
    # Індекси колонок (1-indexed)
    col_code = None
    col_name = None
    col_unit = None
    col_qty = None
    col_price = None
    col_total = None

    # Пошук заголовку таблиці
    for r_idx in range(1, 100):  # Шукаємо заголовок у перших 100 рядках
        row_vals = [sheet.cell(row=r_idx, column=c_idx).value for c_idx in range(1, 15)]
        row_str = [str(val).lower() if val is not None else "" for val in row_vals]

        # Шукаємо ключові слова для визначення стовпчиків
        has_code = any("шифр" in s or "код" in s for s in row_str)
        has_name = any("найменування" in s or "назва" in s or "ресурс" in s for s in row_str)
        has_qty = any("кількість" in s or "об'єм" in s or "кол-во" in s for s in row_str)
        has_price = any("ціна" in s or "цена" in s or "вартість одиниці" in s for s in row_str)

        if (has_code or has_name) and has_qty and has_price:
            header_found = True
            logger.info(f"🔑 Знайдено заголовок таблиці на рядку {r_idx}")
            
            # Визначаємо індекси колонок
            for c_idx, val in enumerate(row_str, 1):
                if not val:
                    continue
                if "шифр" in val or "код" in val:
                    col_code = c_idx
                elif "найменування" in val or "назва" in val or "ресурс" in val:
                    col_name = c_idx
                elif "одиниця" in val or "од. вим." in val or "изм." in val:
                    col_unit = c_idx
                elif "кількість" in val or "об'єм" in val or "кол-во" in val:
                    col_qty = c_idx
                elif "ціна" in val or "цена" in val or "вартість одиниці" in val:
                    # Якщо є кілька цін (відпускна, кошторисна), беремо останню (кошторисну)
                    col_price = c_idx
                elif "всього" in val or "сума" in val or "стоимость" in val:
                    col_total = c_idx
            
            # Захисні значення за замовчуванням, якщо якісь колонки не знайшли
            if col_code is None: col_code = 2
            if col_name is None: col_name = 3
            if col_unit is None: col_unit = 4
            if col_qty is None: col_qty = 5
            if col_price is None: col_price = 7  # Зазвичай кошторисна ціна в АВК йде 7-ю
            if col_total is None: col_total = 8
            
            logger.info(f"Индекси колонок: Код={col_code}, Назва={col_name}, Од={col_unit}, Кіл={col_qty}, Ціна={col_price}, Всього={col_total}")
            start_row = r_idx + 1
            break
    else:
        # Якщо заголовок не знайдено, використовуємо стандартні позиції АВК-5 відомості ресурсів
        logger.warning("⚠️ Не вдалося детерміновано знайти заголовок таблиці. Використовуються дефолтні індекси АВК-5.")
        col_code = 2
        col_name = 3
        col_unit = 4
        col_qty = 5
        col_price = 7
        col_total = 8
        start_row = 5
        header_found = True

    # Збір даних
    current_section = "materials"  # За замовчуванням вважаємо, що спочатку йдуть матеріали
    
    for r_idx in range(start_row, 5000):  # обробляємо до 5000 рядків
        code_val = sheet.cell(row=r_idx, column=col_code).value
        name_val = sheet.cell(row=r_idx, column=col_name).value
        
        # Перевірка на завершення таблиці або кінець листа
        if code_val is None and name_val is None:
            # Спробуємо глянути ще 3 рядки вперед, щоб переконатися, що це не випадкова пуста лінія
            empty_streak = True
            for next_r in range(r_idx + 1, r_idx + 4):
                if sheet.cell(row=next_r, column=col_code).value or sheet.cell(row=next_r, column=col_name).value:
                    empty_streak = False
                    break
            if empty_streak:
                break
            else:
                continue

        code_str = str(code_val).strip() if code_val is not None else ""
        name_str = str(name_val).strip() if name_val is not None else ""
        
        # Визначення зміни розділів (наприклад, перехід до будівельних машин чи трудових ресурсів)
        name_lower = name_str.lower()
        if "будівельні машини" in name_lower or "машини та механізми" in name_lower or "раздел ii" in name_lower:
            current_section = "machinery"
            continue
        elif "трудові ресурси" in name_lower or "витрати труда" in name_lower or "раздел iii" in name_lower:
            current_section = "labor"
            continue

        # Збираємо матеріали або машини
        if current_section not in ("materials", "machinery"):
            continue

        # Ігноруємо підсумкові рядки
        if "всього" in name_lower or "разом" in name_lower or "итого" in name_lower:
            continue

        # Зчитуємо кількість та ціну
        qty_val = sheet.cell(row=r_idx, column=col_qty).value
        price_val = sheet.cell(row=r_idx, column=col_price).value
        total_val = sheet.cell(row=r_idx, column=col_total).value

        # Конвертуємо у числа
        try:
            quantity = float(qty_val) if qty_val is not None else 0.0
            unit_price = float(price_val) if price_val is not None else 0.0
            total_price = float(total_val) if total_val is not None else (quantity * unit_price)
        except (ValueError, TypeError):
            # Якщо це не числа (наприклад, текстові заголовки підрозділів), ігноруємо рядок
            continue

        # Матеріал або машина повинні мати назву та ненульову кількість
        if name_str and quantity > 0:
            item_data = {
                "code": code_str,
                "item": name_str,
                "unit": str(sheet.cell(row=r_idx, column=col_unit).value or "шт").strip(),
                "quantity": quantity,
                "unit_price": unit_price,
                "total": total_price
            }
            if current_section == "materials":
                materials.append(item_data)
            else:
                machinery.append(item_data)

    wb.close()
    logger.info(f"✅ Успішно розпарсено кошторис: знайдено {len(materials)} матеріальних ресурсів та {len(machinery)} машин")
    
    if return_dict:
        return {"materials": materials, "machinery": machinery}
    return materials
