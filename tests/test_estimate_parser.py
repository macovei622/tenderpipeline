"""
tests/test_estimate_parser.py — Тестування парсингу Excel відомості ресурсів
"""
import os
import openpyxl
from ai.estimate_parser import parse_avk_excel_estimate


def test_parse_avk_excel_estimate():
    # 1. Створюємо тимчасовий Excel файл, що імітує вивантаження з АВК-5
    file_path = "temp_test_estimate.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Відомість ресурсів"

    # Пишемо заголовок
    ws.cell(row=1, column=1, value="ПІДСУМКОВА ВІДОМІСТЬ РЕСУРСІВ (імітація)")
    
    # Таблична шапка на рядку 3
    ws.cell(row=3, column=1, value="№")
    ws.cell(row=3, column=2, value="Шифр ресурсу")
    ws.cell(row=3, column=3, value="Найменування ресурсу")
    ws.cell(row=3, column=4, value="Одиниця виміру")
    ws.cell(row=3, column=5, value="Кількість")
    ws.cell(row=3, column=6, value="Відпускна ціна")
    ws.cell(row=3, column=7, value="Кошторисна ціна")
    ws.cell(row=3, column=8, value="Кошторисна вартість, грн")

    # Пишемо розділ будівельних матеріалів
    ws.cell(row=4, column=3, value="I. БУДІВЕЛЬНІ МАТЕРІАЛИ")

    # Рядок з матеріалом 1
    ws.cell(row=5, column=1, value=1)
    ws.cell(row=5, column=2, value="С111-12")
    ws.cell(row=5, column=3, value="Пісок будівельний")
    ws.cell(row=5, column=4, value="м3")
    ws.cell(row=5, column=5, value=100.0)
    ws.cell(row=5, column=6, value=250.0)
    ws.cell(row=5, column=7, value=300.0)
    ws.cell(row=5, column=8, value=30000.0)

    # Рядок з матеріалом 2
    ws.cell(row=6, column=1, value=2)
    ws.cell(row=6, column=2, value="С311-45")
    ws.cell(row=6, column=3, value="Портландцемент М400")
    ws.cell(row=6, column=4, value="т")
    ws.cell(row=6, column=5, value=15.5)
    ws.cell(row=6, column=6, value=4000.0)
    ws.cell(row=6, column=7, value=4500.0)
    ws.cell(row=6, column=8, value=69750.0)

    # Рядок з підсумком (має ігноруватися)
    ws.cell(row=7, column=3, value="Разом матеріальних ресурсів")
    ws.cell(row=7, column=8, value=99750.0)

    # Перехід до іншого розділу (має ігноруватися)
    ws.cell(row=8, column=3, value="II. БУДІВЕЛЬНІ МАШИНИ ТА МЕХАНІЗМИ")
    ws.cell(row=9, column=1, value=3)
    ws.cell(row=9, column=2, value="М200-11")
    ws.cell(row=9, column=3, value="Кран автомобільний")
    ws.cell(row=9, column=4, value="маш-год")
    ws.cell(row=9, column=5, value=8.0)
    ws.cell(row=9, column=7, value=1200.0)
    ws.cell(row=9, column=8, value=9600.0)

    wb.save(file_path)
    wb.close()

    # 2. Запускаємо парсер
    try:
        materials = parse_avk_excel_estimate(file_path)
        
        # 3. Перевіряємо результати
        assert len(materials) == 2, f"Очікувалось 2 матеріали, отримано {len(materials)}"
        
        # Матеріал 1
        assert materials[0]["code"] == "С111-12"
        assert materials[0]["item"] == "Пісок будівельний"
        assert materials[0]["quantity"] == 100.0
        assert materials[0]["unit_price"] == 300.0  # Кошторисна ціна з колонки 7
        assert materials[0]["total"] == 30000.0

        # Матеріал 2
        assert materials[1]["code"] == "С311-45"
        assert materials[1]["item"] == "Портландцемент М400"
        assert materials[1]["quantity"] == 15.5
        assert materials[1]["unit_price"] == 4500.0
        assert materials[1]["total"] == 69750.0

        print("✅ Тест парсера кошторису пройшов УСПІШНО!")
        
    finally:
        # Видаляємо тимчасовий файл
        if os.path.exists(file_path):
            os.remove(file_path)


if __name__ == "__main__":
    test_parse_avk_excel_estimate()
